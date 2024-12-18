from django.http import HttpResponse
from .optimizador import optimizar_turnos
from .models import Guardia, Sucursal, AsignacionNotificacion
import pulp
from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from .forms import GuardiaRegistroForm, SucursalForm
from .utils import calcular_distancia
import pulp
from django.views.generic.edit import UpdateView, DeleteView
from django.urls import reverse_lazy
import pandas as pd
import math
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils import timezone
import openpyxl

def es_supervisor(user):
    return user.groups.filter(name='supervisores').exists()

def user_login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('home')  # Redirigir a una página segura
        else:
            return render(request, 'login/login.html')
    return render(request, 'login/login.html')

def logout_view(request):
    logout(request)
    return redirect('login')  # Redirige a la página de login después del logout

@login_required
def home(request):
    return render(request, 'home/home.html')

def registrar_guardia(request):
    if request.method == 'POST':
        form = GuardiaRegistroForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('registrar_guardia')  # Redirige al login después del registro
    else:
        form = GuardiaRegistroForm()
    return render(request, 'usuarios/registrar_guardia.html', {'form': form})

def ver_usuarios(request):
    # Filtrar guardias con usuarios válidos
    guardias = Guardia.objects.filter(user__isnull=False)
    return render(request, 'usuarios/ver_usuarios.html', {'guardias': guardias})

def exportar_usuarios_excel(request):
    # Crear un libro de trabajo
    workbook = openpyxl.Workbook()
    hoja = workbook.active
    hoja.title = "Usuarios Registrados"

    # Escribir los encabezados
    encabezados = ["ID", "Nombre de Usuario", "Dirección", "Preferencia de Turno", "Cargo"]
    hoja.append(encabezados)

    # Obtener datos de los usuarios registrados
    guardias = Guardia.objects.select_related('user').all()

    for guardia in guardias:
        hoja.append([
            guardia.user.id,
            guardia.user.username,
            guardia.direccion,
            "Mañana" if guardia.preferencia_turno == 1 else "Tarde" if guardia.preferencia_turno == 2 else "Noche",
            guardia.user.groups.all()[0].name if guardia.user.groups.exists() else "Sin grupo",
        ])

    # Configurar la respuesta HTTP para enviar el archivo Excel
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=usuarios_registrados.xlsx"
    workbook.save(response)
    return response

class EditarGuardiaView(UpdateView):
    model = Guardia
    fields = ['direccion', 'latitud', 'longitud', 'disponibilidad_horas', 'activo']
    template_name = 'usuarios/editar_guardia.html'
    success_url = reverse_lazy('ver_usuarios')

class EliminarGuardiaView(DeleteView):
    model = Guardia
    template_name = 'usuarios/confirmar_eliminar_guardia.html'
    success_url = reverse_lazy('ver_usuarios')

@login_required
def registrar_sucursal(request):
    if request.method == 'POST':
        form = SucursalForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('home')
    else:
        form = SucursalForm()
    return render(request, 'sucursales/registrar_sucursal.html', {'form': form})

def ver_sucursales(request):
    # Filtrar sucursales con ID válidos
    sucursales = Sucursal.objects.exclude(id__isnull=True)
    return render(request, 'sucursales/ver_sucursales.html', {'sucursales': sucursales})

class EditarSucursalView(UpdateView):
    model = Sucursal
    fields = ['nombre', 'direccion', 'latitud', 'longitud', 'requerimiento_turnos']
    template_name = 'sucursales/editar_sucursal.html'
    success_url = reverse_lazy('ver_sucursales')  # Redirige después de actualizar

class EliminarSucursalView(DeleteView):
    model = Sucursal
    template_name = 'sucursales/confirmar_eliminar_sucursal.html'
    success_url = reverse_lazy('ver_sucursales')  # Redirige después de eliminar

def exportar_sucursales_excel(request):
    # Crear el libro de trabajo y la hoja
    workbook = openpyxl.Workbook()
    hoja = workbook.active
    hoja.title = "Lista de Sucursales"

    # Agregar encabezados
    encabezados = ["ID", "Nombre", "Dirección", "Latitud", "Longitud", "Requerimiento de Turnos"]
    hoja.append(encabezados)

    # Obtener datos de las sucursales
    sucursales = Sucursal.objects.all()

    # Llenar filas con los datos
    for sucursal in sucursales:
        hoja.append([
            sucursal.id,
            sucursal.nombre,
            sucursal.direccion,
            sucursal.latitud,
            sucursal.longitud,
            sucursal.requerimiento_turnos
        ])

    # Configurar la respuesta HTTP para descargar el archivo
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=sucursales.xlsx"

    # Guardar el libro de trabajo en la respuesta
    workbook.save(response)
    return response

def optimizar_recursos_humanos(request):
    # Obtener datos de la base de datos
    guardias = list(Guardia.objects.all())
    sucursales = list(Sucursal.objects.all())

    # Crear el modelo de optimización
    modelo = pulp.LpProblem("Optimización_Recursos_Humanos", pulp.LpMinimize)

    # Variables de decisión: x[g.id, s.id, t] -> 1 si el guardia g va a la sucursal s en el turno t
    turnos = [1, 2, 3]  # Turnos predefinidos: mañana, tarde, noche
    x = pulp.LpVariable.dicts("x", ((g.id, s.id, t) for g in guardias for s in sucursales for t in turnos), cat="Binary")

    # Parámetros de preferencias de turnos (ejemplo: 1 -> alta preferencia, 3 -> baja preferencia)
    preferencias = {1: 1, 2: 2, 3: 3}  # Menor valor, mayor preferencia

    # Función objetivo: Minimizar distancia + penalización por tiempo y preferencias
    modelo += pulp.lpSum(
        (calcular_distancia(g.latitud, g.longitud, s.latitud, s.longitud) * 0.5 + preferencias[t] * 10) * x[g.id, s.id, t]
        for g in guardias for s in sucursales for t in turnos
    )

    # Restricción 1: Cada sucursal debe cubrir al menos un turno
    for s in sucursales:
        for t in turnos:
            modelo += pulp.lpSum(x[g.id, s.id, t] for g in guardias) >= 1

    # Restricción 2: Un guardia puede estar asignado a un solo turno a la vez
    for g in guardias:
        modelo += pulp.lpSum(x[g.id, s.id, t] for s in sucursales for t in turnos) <= 1

    # Resolver el modelo
    modelo.solve()

    # Extraer resultados únicos
    asignaciones = []
    asignados_guardias = set()

    for g in guardias:
        for s in sucursales:
            for t in turnos:
                if pulp.value(x[g.id, s.id, t]) == 1 and g.id not in asignados_guardias:
                    asignaciones.append((g.user.username, s.nombre, f"Turno {t}"))
                    asignados_guardias.add(g.id)

    # Renderizar resultados
    return render(request, 'resultado_optimizado.html', {'asignaciones': asignaciones})

def calcular_distancia(lat1, lon1, lat2, lon2):
    if not (lat1 and lon1 and lat2 and lon2):
        return float('inf')  # Retorna un valor muy alto si faltan coordenadas
    return math.sqrt((lat2 - lat1) ** 2 + (lon2 - lon1) ** 2) * 100  # Ejemplo simplificado


def calcular_guardias_cercanos(request):
    sucursales = Sucursal.objects.all()  # Recupera todas las sucursales
    sucursal_seleccionada = None
    guardias_cercanos = []
    requerimiento_turnos = 0

    if 'sucursal' in request.GET:
        sucursal_id = request.GET['sucursal']
        if sucursal_id:
            # Obtener sucursal seleccionada
            sucursal_seleccionada = Sucursal.objects.get(id=sucursal_id)
            requerimiento_turnos = sucursal_seleccionada.requerimiento_turnos

            # Obtener guardias activos con coordenadas válidas
            guardias = Guardia.objects.filter(activo=True, latitud__isnull=False, longitud__isnull=False)

            # Calcular distancia y tiempo para cada guardia
            for guardia in guardias:
                distancia = calcular_distancia(
                    sucursal_seleccionada.latitud, sucursal_seleccionada.longitud,
                    guardia.latitud, guardia.longitud
                )
                tiempo = distancia / 50 * 60  # Suponiendo velocidad promedio 50 km/h

                guardias_cercanos.append({
                    'id': guardia.user.id,  # Asegúrate de incluir el ID del usuario aquí
                    'nombre': guardia.user.username,
                    'direccion': guardia.direccion,
                    'distancia': round(distancia, 2),
                    'tiempo': round(tiempo, 2),
                })

            # Ordenar guardias por distancia
            guardias_cercanos.sort(key=lambda x: x['distancia'])

    return render(request, 'sucursales/calcular_guardias_cercanos.html', {
        'sucursales': sucursales,
        'sucursal': sucursal_seleccionada,
        'requerimiento_turnos': requerimiento_turnos,
        'guardias_cercanos': guardias_cercanos,
    })

def exportar_usuarios_asignados(request, sucursal_id):
    # Obtener la sucursal seleccionada
    sucursal = Sucursal.objects.get(id=sucursal_id)

    # Crear un libro de trabajo
    workbook = openpyxl.Workbook()
    hoja = workbook.active
    hoja.title = f"Usuarios Asignados - {sucursal.nombre}"

    # Encabezados
    encabezados = ["ID Guardia", "Nombre Guardia", "Turno", "Fecha de Notificación"]
    hoja.append(encabezados)

    # Obtener notificaciones aceptadas para esta sucursal
    notificaciones = AsignacionNotificacion.objects.filter(
        sucursal=sucursal, estado="aceptada"
    ).select_related("guardia")

    # Llenar los datos en el archivo Excel
    for notificacion in notificaciones:
        hoja.append([
            notificacion.guardia.id,
            notificacion.guardia.username,
            "Mañana" if notificacion.turno == 1 else "Tarde" if notificacion.turno == 2 else "Noche",
            notificacion.fecha_envio.strftime("%Y-%m-%d %H:%M:%S"),
        ])

    # Configurar la respuesta HTTP para descargar el archivo
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=usuarios_asignados_{sucursal.nombre}.xlsx"

    workbook.save(response)
    return response

def enviar_notificacion(request, guardia_id, sucursal_id, turno):
    print(f"ID recibido: {guardia_id}")  # Depuración del ID del usuario
    guardia = User.objects.get(id=guardia_id)
    print(f"Usuario seleccionado: {guardia.id} - {guardia.username}")
    
    sucursal = Sucursal.objects.get(id=sucursal_id)
    AsignacionNotificacion.objects.create(
        guardia=guardia,
        sucursal=sucursal,
        turno=turno,
        estado='pendiente'
    )
    notificacion = AsignacionNotificacion.objects.create(
        guardia=guardia,
        sucursal=sucursal,
        turno=turno,
        estado='pendiente',
        fecha_envio=timezone.now()
        )
    return render(request, 'notificaciones/notificacion_enviada.html', {
        'guardia': guardia,
        'sucursal': sucursal,
        'turno': turno,
        'fecha_envio': notificacion.fecha_envio
    })

@login_required
def ver_notificaciones(request):
    print(f"Usuario autenticado: {request.user.id}, {request.user.username}")

    notificaciones = AsignacionNotificacion.objects.filter(
        guardia_id=request.user, estado='pendiente'
    )

    print("Notificaciones encontradas:", notificaciones)
    return render(request, 'notificaciones/ver_notificaciones.html', {
        'notificaciones': notificaciones
    })


@login_required
def responder_notificacion(request, notificacion_id, respuesta):
    # Obtener la notificación
    notificacion = get_object_or_404(AsignacionNotificacion, id=notificacion_id, guardia=request.user)

    # Actualizar estado según la respuesta
    if respuesta == "aceptar":
        notificacion.estado = "aceptada"
        # Aquí puedes asignar el guardia a la sucursal y turno de forma efectiva
    elif respuesta == "rechazar":
        notificacion.estado = "rechazada"

    notificacion.save()
    return redirect('ver_notificaciones')

@login_required
@user_passes_test(es_supervisor)
def ver_asignaciones_aprobadas(request):
    # Filtrar asignaciones aceptadas
    asignaciones = AsignacionNotificacion.objects.filter(estado='aceptada')
    
    return render(request, 'asignaciones/ver_asignaciones_aprobadas.html', {'asignaciones': asignaciones})

def ver_usuarios_asignados(request, sucursal_id):
    # Obtener la sucursal seleccionada
    sucursal = get_object_or_404(Sucursal, id=sucursal_id)

    # Filtrar notificaciones aceptadas para esta sucursal
    notificaciones = AsignacionNotificacion.objects.filter(
        sucursal=sucursal, estado='aceptada'
    ).select_related('guardia')

    return render(request, 'asignaciones/ver_usuarios_asignados.html', {
        'sucursal': sucursal,
        'notificaciones': notificaciones,
    })

def desasignar_usuario(request, notificacion_id):
    # Obtener la notificación específica
    notificacion = get_object_or_404(AsignacionNotificacion, id=notificacion_id)

    # Eliminar la notificación (o marcarla como rechazada)
    notificacion.delete()

    # Redirigir de vuelta a la página de usuarios asignados
    return redirect('ver_usuarios_asignados', sucursal_id=notificacion.sucursal.id)

def es_guardia(user):
    return user.groups.filter(name="guardias").exists()

@login_required
@user_passes_test(es_guardia)
def ver_sucursales_asignadas(request):
    asignaciones = AsignacionNotificacion.objects.filter(
        guardia=request.user, estado="aceptada"
    ).select_related("sucursal")

    return render(request, "sucursales/sucursales_asignadas.html", {
        "asignaciones": asignaciones
    })